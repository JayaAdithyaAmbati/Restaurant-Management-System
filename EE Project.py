from tkinter import *
from tkinter import messagebox
from tkinter import filedialog
from PIL import ImageTk, Image
from openpyxl import load_workbook as lw
from random import randint as ri
from datetime import datetime


def username(*event):
    if username_entry.get() == 'Username':
        username_entry.delete(0, END)
        username_entry.focus()
    if password_entry.get() != 'Password':
        login_button.config(state=NORMAL)


def password(*event):
    if password_entry.get() == 'Password':
        password_entry.delete(0, END)
    password_entry.config(show='*')
    password_entry.focus()
    if username_entry.get() != 'Username':
        login_button.config(state=NORMAL)


def tab2(*event):
    credentials = lw('RMS Credentials.xlsx')
    credentials = credentials['Credentials']
    for i in range(2, 6):
        if username_entry.get() == str(credentials.cell(i, 2).value):
            if password_entry.get() == str(credentials.cell(i, 3).value):
                login_button.focus_set()
                mainframe1.pack_forget()
                mainframe2.pack()
                t2.set('')

            else:
                messagebox.showerror('ERROR', 'Invalid Password')
            break
    else:
        messagebox.showerror('ERROR', 'Invalid Username')


def total(*event):
    global menu_list
    s = 0
    for category, quantity in menu_list.items():
        sheet = Menu_Data[category]
        for i in range(6):
            c = quantity[i].get()
            if c == '':
                continue
            for j in c:  # To eliminate spaces before first alphanumeric character(if any)
                if j == ' ':
                    quantity[i].delete(0, 1)
                else:
                    break
            c = quantity[i].get()
            if not c.isnumeric():
                messagebox.showerror('ERROR', 'Invalid input(s)')
                return
            elif int(c) == 0:
                continue
            elif int(c) > 100:
                messagebox.showinfo('NOTICE', 'Item limit(100) exceeded')
                return
            s += float(c) * int(sheet.cell(i + 2, 2).value[:-2])
            k = 0
            for j in c:  # To eliminate zeroes before first digit(if any)
                if j == '0':
                    quantity[i].delete(0, 1)
                else:
                    break

    if s == 0:
        messagebox.showerror('ERROR', 'No item selected')
        return

    subtotal_entry.config(state=NORMAL)
    serviceTax_entry.config(state=NORMAL)
    total_entry.config(state=NORMAL)

    subtotal_entry.delete(0, END)
    serviceTax_entry.delete(0, END)
    total_entry.delete(0, END)

    subtotal_entry.insert(0, '{:>7.2f}'.format(s))
    serviceTax_entry.insert(0, '{:>7.2f}'.format(5 * s / 100))
    total_entry.insert(0, '{:>7.2f}'.format(s + 5 * s / 100))

    subtotal_entry.config(state=DISABLED)
    serviceTax_entry.config(state=DISABLED)
    total_entry.config(state=DISABLED)

    subtotal_label.focus_set()


def invoice():
    global menu_list
    total()  # estimate total before invoice (if not totaled before)
    if subtotal_entry.get() == '':
        return
    receipt.config(state=NORMAL)
    receipt.delete('1.0', END)

    bill = ri(1000, 10000)
    date = datetime.now().strftime('%d/%m/%Y')
    time = datetime.now().strftime('%H:%M:%S')

    receipt.insert(END,
                   'BILL:{}\t\t{}\t\t{}\n'.format(str(bill).ljust(13, ' '), date.center(9, ' '), time.center(9, ' ')))
    receipt.insert(END, '*' * 70)
    receipt.insert(END, '\n')
    receipt.insert(END,
                   '{}\t\t{}\t\t{}\n'.format('ITEMS'.ljust(18, ' '), 'QUANTITY'.center(9, ' '), 'PRICE'.center(9, ' ')))
    receipt.insert(END, '*' * 70)
    receipt.insert(END, '\n')
    s = 0
    for category, quantity in menu_list.items():
        sheet = Menu_Data[category]
        for i in range(6):
            c = quantity[i].get()
            if c == '':
                continue
            if int(c) != 0:
                t = '{:>7.2f}'.format(float(c) * int(sheet.cell(i + 2, 2).value[:-2]))
                receipt.insert(END, '{}\t\t{}\t\t{}\n'.format(sheet.cell(i + 2, 1).value.ljust(18, ' '),
                                                              str(c).center(9, ' '),
                                                              ('Rs. ' + str(t)).ljust(9, ' ')))
    receipt.insert(END, '*' * 70)
    receipt.insert(END, '\n{}\t\t{}\t\t{}'.format('Sub Total:'.ljust(18, ' '), str(' ').center(9, ' '),
                                                  ('Rs. ' + subtotal_entry.get()).center(9, ' ')))
    receipt.insert(END, '\n{}\t\t{}\t\t{}'.format('Service Tax:'.ljust(18, ' '), str(' ').center(9, ' '),
                                                  ('Rs. ' + serviceTax_entry.get()).center(9, ' ')))
    receipt.insert(END, '\n{}\t\t{}\t\t{}\n'.format('Total:'.ljust(18, ' '), str(' ').center(9, ' '),
                                                    ('Rs. ' + total_entry.get()).center(9, ' ')))
    receipt.insert(END, '*' * 70)
    receipt.config(state=DISABLED)
    button_save.config(state=NORMAL)


def save():
    if len(receipt.get('1.0', 'end-1c')) == 0:
        return
    file = filedialog.asksaveasfilename(title='Save Invoice', defaultextension='.txt')
    try:
        file = open(file, 'w')
        file.write(receipt.get('1.0', END))
        file.close()
        messagebox.showinfo('Information', 'Your Invoice is saved successfully')
    except FileNotFoundError:
        pass


def reset(*event):
    for i in menu_list.values():
        for j in i:
            j.delete(0, END)
    subtotal_label.focus_set()

    subtotal_entry.config(state=NORMAL)
    serviceTax_entry.config(state=NORMAL)
    total_entry.config(state=NORMAL)
    receipt.config(state=NORMAL)

    subtotal_entry.delete(0, END)
    serviceTax_entry.delete(0, END)
    total_entry.delete(0, END)
    receipt.delete('1.0', END)

    subtotal_entry.config(state=DISABLED)
    serviceTax_entry.config(state=DISABLED)
    total_entry.config(state=DISABLED)
    receipt.config(state=DISABLED)

    button_save.config(state=DISABLED)


def logout():
    if messagebox.askyesno('LOGOUT', 'Are you sure you want to Logout ?'):
        reset()
        close_calci()
        mainframe2.pack_forget()
        mainframe1.pack()
        password_entry.focus()



def calculator():
    if calci.wm_state() == 'withdrawn' or calci.wm_state() == 'iconic':
        calci.deiconify()
    calci.focus_force()


def close_calci():
    calci.withdraw()
    clear()


def close_root():
    if mainframe2.winfo_ismapped():
        if messagebox.askyesno('QUIT', 'Are you sure you want to Quit ?'):
            calci.destroy()
            root.destroy()
    else:
        calci.destroy()
        root.destroy()



def clear():
    Disp_box.delete(0, END)


def backspace():
    global flag, exp
    s = Disp_box.get()
    clear()
    if flag == 0:
        Disp_box.insert(0, s[:-1])
    else:
        Disp_box.insert(0, exp[:-1])
    flag = 0


def click_number(n):
    global flag
    if flag == 2 or (flag == 1 and n not in operators):
        clear()
    s = Disp_box.get()
    if (len(s) == 1 and s[-1] == '0') or (len(s) > 1 and (s[-1] == '0' and s[-2] in operators)):
        backspace()
    flag = 0
    Disp_box.insert(END, str(n))


def click_operator(n):
    global flag
    if flag == 2:
        clear()
    s = Disp_box.get()
    if (len(s) == 0 and n != '-') or (len(s) == 1 and s[-1] == '-'):
        return
    elif s != '':
        if n != '-':
            if s[-1] in operators:
                backspace()
                if s[-2] in operators:
                    backspace()
        else:
            if s[-1] in operators[:2]:
                backspace()
    flag = 0
    Disp_box.insert(END, n)


def click_dot(n):
    global flag
    if flag == 2 or (flag == 1 and n not in operators):
        clear()
    s = Disp_box.get()
    if s != '':
        if s[-1] == '.':
            return
        elif s[-1] in operators:
            n = '0' + n
    else:
        n = '0' + n
    Disp_box.insert(END, n)
    flag = 0


def equate(*event):
    global flag, exp
    s = exp = Disp_box.get()
    if len(s) == 0 or (len(s) > 0 and s[-1] in operators):
        return
    s = s.replace('×', '*')
    s = s.replace('÷', '/')
    try:
        s = eval(s)
        flag = 1
        if type(s) == float and s.is_integer():
            s = int(s)
    except ZeroDivisionError:
        s = 'Cannot divide by zero'
        flag = 2
    except (NameError, SyntaxError, TypeError):
        s = 'ERROR'
        flag = 2

    clear()
    Disp_box.insert(0, s)


# Driver code

root = Tk()
root.title('Restaurant Management system')
root.geometry('900x600')
root.state('zoomed')
root.config(bg='#000080')
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

login_image = ImageTk.PhotoImage(
    Image.open('Login pic.jpeg').resize((screen_width, screen_height), Image.Resampling.LANCZOS))
calci_icon = ImageTk.PhotoImage(Image.open('Calculator icon.png').resize((35, 35), Image.Resampling.LANCZOS))
Menu_Data = lw('RMS Menu.xlsx')

mainframe1 = Frame(root)
mainframe1.pack()

login_page = Label(mainframe1, image=login_image, bd=0)
login_page.grid(row=0, column=0)

login_label = Label(mainframe1, text='Log in', font=('Inter', 48, 'bold'), bg='#FFFFFF')
login_label.place(x=550, y=265)

t1 = StringVar()
t2 = StringVar()
t1.set('Username')
t2.set('Password')

username_entry = Entry(mainframe1, font=('Inter', 25), textvariable=t1, bg='#ADADFF', bd=7, width=23, relief=RIDGE)
username_entry.place(x=550, y=390)

password_entry = Entry(mainframe1, font=('Inter', 25), textvariable=t2, bg='#ADADFF', bd=7, width=23, relief=RIDGE,
                       show='')
password_entry.place(x=550, y=460)

login_label.focus_set()
login_label.bind('<Tab>', username)
username_entry.bind('<Button-1>', username)
username_entry.bind('<Tab>', password)
username_entry.bind('<Return>', password)
password_entry.bind('<Button-1>', password)
password_entry.bind('<Return>', tab2)

login_button = Button(mainframe1, font=('Inter', 24), text='Log in', command=tab2, bg='#00A3FF', fg='#FFFFFF', width=22,
                      state=DISABLED)
login_button.place(x=550, y=530)

mainframe2 = Frame(root, bg='#000080')
# mainframe2.pack()

# Frames

Top_Frame = Frame(mainframe2, bg='#000080')
Top_Frame.pack(side=TOP)

title_label = Label(Top_Frame, text='Restaurant Management System', font=('Calibri', 50, 'bold'), bg='#000080',
                    fg='Yellow', width=screen_width // 35, padx=4, bd=12, relief=RIDGE)
title_label.grid(row=0, column=0)

Left_frame = Frame(mainframe2, bd=12, relief=RIDGE, bg='#000080')
Left_frame.pack(side=LEFT)

menu_frame = Frame(Left_frame, bg='#000000')
menu_frame.pack(side=LEFT)

Starters_frame = LabelFrame(menu_frame, text='STARTERS', font=('Calibri', 20, 'bold'), bg='#000000', fg='yellow', bd=10)
Starters_frame.grid(row=0, column=0, padx=3)

Rotis_frame = LabelFrame(menu_frame, text='ROTIS', font=('Calibri', 20, 'bold'), bg='#000000', fg='yellow', bd=10)
Rotis_frame.grid(row=0, column=1, padx=3)

MainCourse_frame = LabelFrame(menu_frame, text='MAIN COURSE', font=('Calibri', 20, 'bold'), bg='#000000', fg='yellow',
                              bd=10)
MainCourse_frame.grid(row=0, column=2, padx=3)

Tiffins_frame = LabelFrame(menu_frame, text='TIFFINS', font=('Calibri', 20, 'bold'), bg='#000000', fg='yellow', bd=10)
Tiffins_frame.grid(row=1, column=0, padx=3, pady=3)

Beverages_frame = LabelFrame(menu_frame, text='BEVERAGES', font=('Calibri', 20, 'bold'), bg='#000000', fg='yellow',
                             bd=10)
Beverages_frame.grid(row=1, column=1, padx=3)

Desserts_frame = LabelFrame(menu_frame, text='DESSERTS', font=('Calibri', 20, 'bold'), bg='#000000', fg='yellow', bd=10)
Desserts_frame.grid(row=1, column=2, padx=3)

Right_outer_frame = Frame(mainframe2, bg='#000080', relief=RIDGE, bd=12)
Right_outer_frame.pack(side=RIGHT)

Right_frame = Frame(Right_outer_frame, bg='DarkOrange', relief=RIDGE)
Right_frame.pack()

receipt_frame = Frame(Right_frame, bd=10, relief=RIDGE)
receipt_frame.pack()

amount_frame = Frame(Right_frame, bg='Gray')
amount_frame.pack(fill=X)

buttons_frame = Frame(Right_frame, bd=10, relief=RIDGE, bg='DarkOrange')
buttons_frame.pack()

# Menu

menu_frame_list = [Starters_frame, Rotis_frame, MainCourse_frame, Tiffins_frame, Beverages_frame, Desserts_frame]

menu_list = {'STARTERS': [], 'ROTIS': [], 'MAIN COURSE': [], 'TIFFINS': [], 'BEVERAGES': [], 'DESSERTS': []}

for i in menu_frame_list:
    text = i.cget('text')
    sheet = Menu_Data[text]
    for j in range(2, sheet.max_row+1):
        item_label = Label(i, text=sheet.cell(j, 1).value.ljust(19, ' '), font=('Calibri', 15), bg='#000000',
                           fg='Yellow')
        item_label.grid(row=j, column=0, sticky=W, padx=3)
        rate_label = Label(i, text=sheet.cell(j, 2).value.rjust(5, ' '), font=('Calibri', 15), bg='#000000',
                           fg='Yellow', width=screen_width // 307)
        rate_label.grid(row=j, column=1, sticky=E, padx=3, pady=screen_height // 87)
        entry_box = Entry(i, font=('Arial', 15), bd=7, width=4, justify=RIGHT, bg='#ADADFF', fg='red')
        entry_box.grid(row=j, column=2, padx=3)
        entry_box.bind('<Return>', total)
        menu_list[text].append(entry_box)
else:
    entry_box.bind('<Tab>', total)

# Right Part

y = Scrollbar(receipt_frame, orient='vertical')
y.grid(row=0, column=1, sticky=N + S)

receipt = Text(receipt_frame, font=('Helvatica', 14), width=screen_width // 36, height=screen_height // 52, bd=10,
               yscrollcommand=y.set, state=DISABLED)
receipt.grid(row=0, column=0)

y.config(command=receipt.yview)

subtotal_label = Label(amount_frame, text='Sub Total ', font=('Calibri', 20, 'bold'), bg='Gray', fg='#000000')
subtotal_label.grid(row=0, column=0, padx=20, pady=14)

serviceTax_label = Label(amount_frame, text='Service Tax ', font=('Calibri', 20, 'bold'), bg='Gray', fg='#000000')
serviceTax_label.grid(row=1, column=0, padx=20, pady=14)

total_label = Label(amount_frame, text='Total ', font=('Calibri', 20, 'bold'), bg='Gray', fg='#000000')
total_label.grid(row=2, column=0, padx=20, pady=13)

subtotal_entry = Entry(amount_frame, font=('Calibri', 20, 'bold'), state=DISABLED, justify=RIGHT, bd=5, relief=RIDGE,
                       bg='#F0F0F8')
subtotal_entry.grid(row=0, column=1)

serviceTax_entry = Entry(amount_frame, font=('Calibri', 20, 'bold'), state=DISABLED, justify=RIGHT, bd=5, relief=RIDGE,
                         bg='#F0F0F8')
serviceTax_entry.grid(row=1, column=1)

total_entry = Entry(amount_frame, font=('Calibri', 20, 'bold'), state=DISABLED, justify=RIGHT, bd=5, relief=RIDGE,
                    bg='#F0F0F8')
total_entry.grid(row=2, column=1)

calculator_button = Button(amount_frame, image=calci_icon, command=calculator)
calculator_button.grid(row=2, column=2, rowspan=2, padx=3)

button_total = Button(buttons_frame, text='Total', font=('Calibri', 18, 'bold'), bg='DarkOrange', fg='#000000', bd=5,
                      padx=9, command=total)
button_total.grid(row=0, column=0)

button_invoice = Button(buttons_frame, text='Invoice', font=('Calibri', 18, 'bold'), bg='DarkOrange', fg='#000000',
                        bd=5, padx=8, command=invoice)
button_invoice.grid(row=0, column=1)

button_save = Button(buttons_frame, text='Save', font=('Calibri', 18, 'bold'), bg='DarkOrange', fg='#000000', bd=5,
                     padx=10, command=save, state=DISABLED)
button_save.grid(row=0, column=2)

button_reset = Button(buttons_frame, text='Reset', font=('Calibri', 18, 'bold'), bg='DarkOrange', fg='#000000', bd=5,
                      padx=9, command=reset)
button_reset.grid(row=0, column=3)

button_logout = Button(buttons_frame, text='Logout', font=('Calibri', 18, 'bold'), bg='DarkOrange', fg='#000000', bd=5,
                       padx=7, command=logout)
button_logout.grid(row=0, column=4)

# Calculator Window

calci = Tk()
calci.resizable(0, 0)
calci.title('Calculator')
calci.geometry('392x380+590+250')
operators = '-+×÷'
exp = ''
flag = 0

Disp_box = Entry(calci, width=24, borderwidth=15, justify=RIGHT, bg='#FFFFC1', font=('Inter', 20))
Disp_box.grid(row=0, column=0, columnspan=4)

button_clear = Button(calci, text='CLEAR', width=24, padx=11, height=2, font=('Inter', 15), bg='gray',
                      activebackground='gray', command=clear)
button_backspace = Button(calci, text='DEL', width=8, height=2, font=('Inter', 15), bg='gray',
                          activebackground='gray', command=backspace)

button_add = Button(calci, text='+', width=8, pady=13, font=('Inter', 15), bg='gray', activebackground='gray',
                    command=lambda: click_operator('+'))
button_sub = Button(calci, text='-', width=8, height=2, font=('Inter', 15), bg='gray', activebackground='gray',
                    command=lambda: click_operator('-'))
button_div = Button(calci, text='÷', width=8, height=2, font=('Inter', 15), bg='gray', activebackground='gray',
                    command=lambda: click_operator('÷'))
button_mul = Button(calci, text='×', width=8, height=2, font=('Inter', 15), bg='gray', activebackground='gray',
                    command=lambda: click_operator('×'))

button_1 = Button(calci, text='1', width=8, height=2, font=('Inter', 15), bg='cyan', activebackground='cyan',
                  command=lambda: click_number('1'))
button_2 = Button(calci, text='2', width=8, height=2, font=('Inter', 15), bg='cyan', activebackground='cyan',
                  command=lambda: click_number('2'))
button_3 = Button(calci, text='3', width=8, height=2, font=('Inter', 15), bg='cyan', activebackground='cyan',
                  command=lambda: click_number('3'))
button_4 = Button(calci, text='4', width=8, height=2, font=('Inter', 15), bg='cyan', activebackground='cyan',
                  command=lambda: click_number('4'))
button_5 = Button(calci, text='5', width=8, height=2, font=('Inter', 15), bg='cyan', activebackground='cyan',
                  command=lambda: click_number('5'))
button_6 = Button(calci, text='6', width=8, height=2, font=('Inter', 15), bg='cyan', activebackground='cyan',
                  command=lambda: click_number('6'))
button_7 = Button(calci, text='7', width=8, height=2, font=('Inter', 15), bg='cyan', activebackground='cyan',
                  command=lambda: click_number('7'))
button_8 = Button(calci, text='8', width=8, height=2, font=('Inter', 15), bg='cyan', activebackground='cyan',
                  command=lambda: click_number('8'))
button_9 = Button(calci, text='9', width=8, height=2, font=('Inter', 15), bg='cyan', activebackground='cyan',
                  command=lambda: click_number('9'))
button_0 = Button(calci, text='0', width=8, pady=13, font=('Inter', 15), bg='cyan', activebackground='cyan',
                  command=lambda: click_number('0'))

button_dot = Button(calci, text='.', width=8, pady=13, font=('Inter', 15), bg='cyan', activebackground='cyan',
                    command=lambda: click_dot('.'))
button_equal = Button(calci, text='=', width=8, pady=13, font=('Inter', 15), bg='blue', activebackground='blue',
                      command=equate)

button_clear.grid(row=1, column=0, columnspan=3)
button_backspace.grid(row=1, column=3)

button_7.grid(row=2, column=0)
button_8.grid(row=2, column=1)
button_9.grid(row=2, column=2)
button_div.grid(row=2, column=3)

button_4.grid(row=3, column=0)
button_5.grid(row=3, column=1)
button_6.grid(row=3, column=2)
button_mul.grid(row=3, column=3)

button_1.grid(row=4, column=0)
button_2.grid(row=4, column=1)
button_3.grid(row=4, column=2)
button_sub.grid(row=4, column=3)

button_dot.grid(row=5, column=0)
button_0.grid(row=5, column=1)
button_equal.grid(row=5, column=2)
button_add.grid(row=5, column=3)

calci.bind('<Return>', equate)
calci.withdraw()

calci.protocol('WM_DELETE_WINDOW', close_calci)
root.protocol('WM_DELETE_WINDOW', close_root)

calci.mainloop()
root.mainloop()
